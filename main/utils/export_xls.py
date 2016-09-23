from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Color, Style, PatternFill
from django.utils.translation import ugettext_lazy as _

from main.models import *
from django.utils import timezone

HEADER = [str(_('Titre')),
          str(_('Auteur(s)')),
          str(_('Proprietaire(s)')),
          str(_('Lu/pas lu'))]


def export_xls_livres(request):
    livres = Livre.find_all()
    return impression(request, livres, 'livres')


def impression(request, livres, titre):
    wb = Workbook()
    ws = wb.active
    ws.title = titre
    __columns_ajusting(ws)
    if request.user.is_authenticated():
        ws.append(HEADER)
        nom_fichier = str(request.user) + "_" + titre + ".xlsx"
    else:
        ws.append([str(_('Titre')),
                   str(_('Auteur(s)')),
                   str(_('Proprietaire(s)'))])
        nom_fichier = "livres.xlsx"

    cptr = 1
    for l in livres:
        proprietaires = l.proprietaires
        proprio = ""
        cpt = 0
        for p in proprietaires:
            if cpt > 0:
                proprio += ", " + p.personne.nom + " " + p.personne.prenom
            else:
                proprio += p.personne.nom + " " + p.personne.prenom
            cpt = cpt + 1
        etat_lecture = ""

        if request.user.is_authenticated():
            lecture = Livre.find_lecture(l.id, request.user)
            if lecture:
                etat_lecture = "Lu"
        auteurs = l.auteurs_livres_str
        ws.append([l.titre,
                   str(auteurs),
                   str(proprio),
                   etat_lecture])
    ws1 = wb.create_sheet()
    ws1.append([timezone.now()])
    if request.user.is_authenticated():
        personne = Personne.find_personne_by_user(request.user)
        val = personne.nom + ", " + personne.prenom
        ws1.append([val])

    ws1.title = 'Paramètres'
    col_date = ws1.column_dimensions['A']
    col_date.width = 40
    response = HttpResponse(content=save_virtual_workbook(wb))
    response['Content-Disposition'] = 'attachment; filename=%s' % nom_fichier
    response['Content-type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response


def export_xls_mes_livres(request):
    livres = Livre.find_all_by_user(request.user)
    return impression(request, livres, 'mes_livres')


def export_xls_livres_lu(request):
    livres = Lecture.find_all_by_user(request.user)

    return impression(request, livres, 'livres_lus')


def __columns_ajusting(ws):
    """
    ajustement des colonnes à la bonne dimension
    """
    col_academic_year = ws.column_dimensions['A']
    col_academic_year.width = 40
    col_academic_year = ws.column_dimensions['B']
    col_academic_year.width = 40
    col_academic_year = ws.column_dimensions['C']
    col_academic_year.width = 20


def export_xls_livres_by_auteur(request):
    auteurs = Auteur.find_all()
    return impression_by_auteur(request, auteurs, 'livres_par_auteurs')


def impression_by_auteur(request, auteurs, titre):
    wb = Workbook()
    ws = wb.active
    ws.title = titre
    __columns_ajusting(ws)
    if request.user.is_authenticated():
        ws.append(HEADER)
        nom_fichier = str(request.user) + "_" + titre + ".xlsx"
    else:
        ws.append([str(_('Auteur(s)')),
                   str(_('Titre')),
                   str(_('Proprietaire(s)'))])
        nom_fichier = "livres.xlsx"

    for auteur in auteurs:
        livres = Livre.find_by_auteur(auteur.id)
        for l in livres:
            proprietaires = l.proprietaires
            proprio = ""
            cpt = 0
            for p in proprietaires:
                if cpt > 0:
                    proprio += ", " + p.personne.nom + " " + p.personne.prenom
                else:
                    proprio += p.personne.nom + " " + p.personne.prenom
                cpt = cpt + 1
            etat_lecture = ""

            if request.user.is_authenticated():
                lecture = Livre.find_lecture(l.id, request.user)
                if lecture:
                    etat_lecture = "Lu"
            auteurs = l.auteurs_livres_str
            ws.append([str(auteurs),
                       l.titre,
                       str(proprio),
                       etat_lecture])
    ws1 = wb.create_sheet()
    ws1.append([timezone.now()])
    if request.user.is_authenticated():
        personne = Personne.find_personne_by_user(request.user)
        val = personne.nom + ", " + personne.prenom
        ws1.append([val])

    ws1.title = 'Paramètres'
    col_date = ws1.column_dimensions['A']
    col_date.width = 40
    response = HttpResponse(content=save_virtual_workbook(wb))
    response['Content-Disposition'] = 'attachment; filename=%s' % nom_fichier
    response['Content-type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response
